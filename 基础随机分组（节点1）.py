import random
from openpyxl import Workbook
import sys
num_people = int(sys.argv[1])  # 获取用户输入的人数
rows = int(sys.argv[2])
cols = int(sys.argv[3])    # 定义矩阵的行数和列数

elements = [str(i) for i in range(1, num_people + 1)]   # 生成一个包含从1到人数的数字序列

wb = Workbook()
ws = wb.active    # 创建一个Excel工作簿

# 遍历矩阵的每个单元格
for row in range(1, rows + 1):     # 遍历行
   for col in range(1, cols + 1):    # 遍历列
        if elements:        # 如果还有元素可以插入，则随机选择一个元素插入
            element = random.choice(elements)
            elements.remove(element)
            ws.cell(row=row, column=col, value=element)
        else:          # 否则，将单元格留空
            ws.cell(row=row, column=col, value='')

wb.save('matrix.xlsx')      # 保存Excel文件

import os
folder_path = os.path.dirname(os.path.abspath('matrix.xlsx'))       # 获取文件所在的文件夹路径
os.startfile(folder_path)       # 打开文件所在的文件夹