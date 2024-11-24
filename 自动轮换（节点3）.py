import openpyxl
import subprocess
import os
import pandas as pd

df = pd.read_excel('matrix.xlsx')  # 读取Excel文件


def last_non_space_row(column):  # 定义一个函数，用于找到最后一行非空或非空格的行号
    for i in range(len(column) - 1, -1, -1):    # 从最后一行开始向前遍历
        if pd.notnull(column[i]) and str(column[i]).strip() != '':   # 如果单元格非空且非空格
            return i + 1  # 返回行号（加1因为行号从1开始）
    return None  # 如果整列都是空或空格，则返回None


last_non_space_rows = df.apply(last_non_space_row, axis=0)  # 应用函数到每一列，并获取结果
in_row = list(map(int, last_non_space_rows+2))  # 将结果转换为整数列表

wb = openpyxl.load_workbook('matrix.xlsx')  # 打开Excel文件
ws = wb.active  # 获取活动工作表
first_row = ws[1]  # 获取第一行的数据
first_row_values = [cell.value for cell in first_row]  # 获取第一行数据的值
for i, value in enumerate(first_row_values, start=1):  # 遍历第一行数据的值
    for x in in_row:  # 遍历最后一行非空或非空格的行号
        ws.cell(row=x, column=i).value = value  # 将第一行数据的值复制到最后一行非空或非空格的行
        if in_row.index(x) == 0:  # 检查x是否为第一个元素
            in_row.remove(x)    # 如果是，则从列表中删除该元素
            break  # 跳出循环
ws.delete_rows(1)    # 删除第一行数据
wb.save('matrix.xlsx')  # 保存修改后的Excel文件

first_col = ws['A']  # 获取第一列数据
data_list = [cell.value for cell in first_col if cell.value is not None]  # 获取第一列非空数据的值
n_col = ws.max_column + 1  # 获取当前最大列数，并加1以添加新数据
for i, value in enumerate(data_list, start=1):  # 遍历第一列非空数据的值
    ws.cell(row=i, column=n_col).value = value  # 将第一列非空数据的值复制到新列
ws.delete_cols(1)   # 删除第一列数据
wb.save('matrix.xlsx')  # 保存修改后的Excel文件

folder_path = os.path.dirname(os.path.abspath('matrix.xlsx'))  # 获取文件所在的文件夹路径
subprocess.run(['start', 'excel.exe', 'matrix.xlsx'], shell=True)  # 使用subprocess模块调用Excel应用程序打开文件
